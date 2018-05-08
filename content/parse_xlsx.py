from openpyxl import load_workbook

from _common import Dictionary, DictionaryEntry


def ParseXLSX(path):
    dictionary = Dictionary()

    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]

    for row in ws.rows:
        if row[0].value:
            entry = DictionaryEntry(str(row[0].value) if row[0].value is not None else None,
                                explanation=str(row[1].value) if row[1].value is not None else None,
                                same=str(row[2].value) if row[2].value is not None else None,
                                close=str(row[3].value) if row[3].value is not None else None,
                                usage=str(row[4].value) if row[4].value is not None else None,
                                construction=str(row[5].value) if row[5].value is not None else None,
                                source=str(row[6].value) if row[6].value is not None else None,
                                commonsource=str(row[7].value) if row[7].value is not None else None)
            dictionary.add(entry)

    return dictionary
